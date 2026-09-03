#!/usr/bin/env python3
"""Generate Hope Anthology directory content from a sibling To Make or To Keep workbook.

The source workbook is the content authority.  Generated JavaScript files must never be
edited by hand; amend the workbook and rerun this script instead.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - clear operational failure
    raise SystemExit("openpyxl is required. Install it with: python3 -m pip install openpyxl") from exc


SCRIPT_PATH = "tools/generate_directory_content.py"
PLACEHOLDER_TEXT = "placeholder"
PALE_YELLOW = "FFF2CC"
URL_RE = re.compile(r"^https://[^\s]+$", re.IGNORECASE)

ASSETS = {
    "logo": "https://images.squarespace-cdn.com/content/6a258894c750534b28845855/1f3bfb43-558e-4262-b2e2-d4e3b56bd77e/01-the-hope-anthology.jpg?content-type=image%2Fjpeg",
    "star": "https://images.squarespace-cdn.com/content/6a258894c750534b28845855/470da432-cdf9-45be-b3ba-2e9599ce5f4f/04-the-hope-anthology-botanical-star.png?content-type=image%2Fpng",
}

NAVIGATION = [
    {"label": "To Keep", "url": "/to-keep"},
    {"label": "To Make", "url": "/to-make"},
    {"label": "The Story", "url": "/story"},
    {"label": "Collaborate", "url": "/for-artists"},
    {"label": "Collective", "url": "/collective"},
]

FOOTER = {
    "privacyUrl": "/privacy-policy",
    "accessibilityUrl": "/accessibility",
    "copyright": "© The Hope Anthology 2026",
}

COLLECTIVE = {
    "heading": "Something worth being part of.",
    "body": "New designs, behind-the-scenes making, and the occasional reminder that you’re doing better than you think. No noise. No pressure. Just the Anthology.",
    "buttonLabel": "Join the Collective",
    "buttonUrl": "/collective",
}

TO_MAKE_ROSTER_HEADERS = {
    "active": ("Active\n(YES/NO)", "Active"),
    "slug": ("Slug",),
    "display_name": ("Display Name",),
    "button_name": ("Button Name",),
    "artist_number": ("Artist Number",),
    "badge": ("Badge",),
    "medium": ("Medium",),
    "collections": ("Collections\n(comma-separated)", "Collections"),
    "price_from": ("Price From\n(e.g. £3.50)", "Price From\n(e.g. £15)", "Price From"),
    "difficulty": ("Difficulty",),
    "technique": ("Technique",),
    "delivery": ("Delivery",),
    "feeling": ("Feeling\n(card tagline)", "Feeling"),
    "trait_1": ("Trait 1",),
    "trait_2": ("Trait 2",),
    "trait_3": ("Trait 3",),
    "card_url": ("Card URL",),
    "hero_image_url": ("Hero Image URL",),
    "hero_image_alt": ("Hero Image Alt",),
    "notes": ("Notes",),
}

TO_MAKE_PATTERN_HEADERS = {
    "active": ("Active\n(YES/NO)", "Active"),
    "collection": ("Collection",),
    "collection_order": ("CollectionOrder", "Collection Order"),
    "product_order": ("ProductOrder", "Product Order"),
    "title": ("Pattern Title", "Product Title", "Title"),
    "status": ("Status\n(available/coming)", "Status"),
    "format": ("Format",),
    "difficulty": ("Difficulty",),
    "technique": ("Technique", "Medium"),
    "delivery": ("Delivery", "Download"),
    "price": ("Price (£)\nnumber only", "Price (£)", "Price"),
    "etsy_url": ("Etsy URL", "Primary URL", "Product URL"),
    "description": ("Meaning /Description", "Meaning / Description", "Description"),
    "good_for_1": ("Good For 1",),
    "good_for_2": ("Good For 2",),
    "good_for_3": ("Good For 3",),
    "workbench_url": ("Image — Workbench URL",),
    "workbench_alt": ("Image — Workbench Alt",),
    "lifestyle_url": ("Image — Lifestyle URL",),
    "lifestyle_alt": ("Image — Lifestyle Alt",),
    "inside_url": ("Image — Inside URL",),
    "inside_alt": ("Image — Inside Alt",),
    "coming_soon_label": ("Coming Soon\nLabel", "Coming Soon Label"),
    "notes": ("Notes",),
}

TO_KEEP_PATTERN_HEADERS = {
    "active": ("Active\n(YES/NO)", "Active(YES/NO)", "Active"),
    "collection": ("Collection",),
    "collection_order": ("CollectionOrder", "Collection Order"),
    "product_order": ("ProductOrder", "Product Order"),
    "title": ("Product Title", "Pattern Title", "Title"),
    "status": ("Status\n(available/coming)", "Status(available/coming)", "Status"),
    "format": ("Format",),
    "sizes": ("Sizes",),
    "price": ("Price (£)\nnumber only", "Price (£)number only", "Price (£)", "Price"),
    "etsy_url": ("Etsy URL", "Primary URL", "Product URL"),
    "description": ("Meaning /Description", "Meaning / Description", "Description"),
    "gift_for_1": ("Gift For 1",),
    "gift_for_2": ("Gift For 2",),
    "gift_for_3": ("Gift For 3",),
    "lifestyle_url": ("Image — Lifestyle URL",),
    "lifestyle_alt": ("Image — Lifestyle Alt",),
    "flat_url": ("Image — Flat URL",),
    "flat_alt": ("Image — Flat Alt",),
    "detail_url": ("Image — Detail URL",),
    "detail_alt": ("Image — Detail Alt",),
    "coming_soon_label": ("Coming Soon\nLabel", "Coming SoonLabel", "Coming Soon Label"),
    "notes": ("Notes",),
}

PROFILE_FIELDS = {
    "slug": ("Slug",),
    "display_name": ("Display Name",),
    "discipline": ("Discipline",),
    "bio": ("Bio",),
    "meta_description": ("Meta Description",),
    "stats_collections": ("Stats — Collections", "Stats - Collections"),
    "stats_patterns": ("Stats — Patterns", "Stats - Patterns"),
    "stats_from_price": ("Stats — From Price", "Stats - From Price"),
    "montage_1_url": ("Montage Image 1 URL",),
    "montage_1_alt": ("Montage Image 1 Alt",),
    "montage_2_url": ("Montage Image 2 URL",),
    "montage_2_alt": ("Montage Image 2 Alt",),
    "montage_3_url": ("Montage Image 3 URL",),
    "montage_3_alt": ("Montage Image 3 Alt",),
}

TO_KEEP_PROFILE_FIELDS = {
    **PROFILE_FIELDS,
    "stats_patterns": ("Stats — Prints", "Stats - Prints", "Stats — Patterns", "Stats - Patterns"),
}

REQUIRED_ROSTER_HEADERS = (
    "active", "slug", "display_name", "button_name", "artist_number", "badge", "medium", "collections", "price_from",
    "difficulty", "technique", "delivery", "feeling", "trait_1", "trait_2", "trait_3",
    "card_url", "hero_image_url", "hero_image_alt",
)
REQUIRED_TO_KEEP_ROSTER_HEADERS = (
    "active", "slug", "display_name", "button_name", "artist_number", "badge", "medium", "collections", "price_from",
    "feeling", "trait_1", "trait_2", "trait_3", "card_url", "hero_image_url", "hero_image_alt",
)
REQUIRED_TO_MAKE_ROSTER_FIELDS = (
    "active", "slug", "display_name", "badge", "medium", "collections", "price_from",
    "difficulty", "technique", "delivery", "feeling", "trait_1", "trait_2", "trait_3",
    "card_url", "hero_image_url", "hero_image_alt",
)
REQUIRED_TO_KEEP_ROSTER_FIELDS = (
    "active", "slug", "display_name", "badge", "medium", "collections", "price_from",
    "feeling", "trait_1", "trait_2", "trait_3", "card_url", "hero_image_url", "hero_image_alt",
)
REQUIRED_TO_MAKE_PROFILE_FIELDS = tuple(PROFILE_FIELDS)
REQUIRED_TO_KEEP_PROFILE_FIELDS = tuple(TO_KEEP_PROFILE_FIELDS)
REQUIRED_TO_MAKE_PATTERN_FIELDS = (
    "active", "collection", "collection_order", "product_order", "title", "status", "format",
    "difficulty", "technique", "delivery", "price", "etsy_url", "description", "workbench_url",
    "workbench_alt", "lifestyle_url", "lifestyle_alt", "inside_url", "inside_alt",
)
REQUIRED_TO_KEEP_PATTERN_FIELDS = (
    "active", "collection", "collection_order", "product_order", "title", "status", "format", "sizes",
    "price", "etsy_url", "description", "lifestyle_url", "lifestyle_alt", "flat_url", "flat_alt",
    "detail_url", "detail_alt",
)


@dataclass(frozen=True)
class Issue:
    severity: str
    sheet: str
    cell: str
    field: str
    message: str


class Audit:
    def __init__(self, mode: str) -> None:
        self.mode = mode
        self.issues: list[Issue] = []
        self.placeholders: list[Issue] = []
        self.roster_records: list[dict[str, Any]] = []
        self.blocked = False

    def add(self, severity: str, sheet: str, cell: str, field: str, message: str) -> None:
        self.issues.append(Issue(severity, sheet, cell, field, message))

    def required_problem(self, sheet: str, cell: str, field: str, message: str) -> None:
        self.add("error" if self.mode == "strict" else "warning", sheet, cell, field, message)

    def warning(self, sheet: str, cell: str, field: str, message: str) -> None:
        """Record a non-blocking staging or editorial warning in either mode."""
        self.add("warning", sheet, cell, field, message)

    def blocking_error(self, sheet: str, cell: str, field: str, message: str) -> None:
        """Record an error that prevents every generated file from being replaced."""
        self.add("error", sheet, cell, field, message)
        self.blocked = True

    def placeholder(self, sheet: str, cell: str, field: str, message: str) -> None:
        issue = Issue("error" if self.mode == "strict" else "warning", sheet, cell, field, message)
        self.placeholders.append(issue)
        self.issues.append(issue)

    @property
    def fatal(self) -> list[Issue]:
        return [issue for issue in self.issues if issue.severity == "fatal"]

    @property
    def errors(self) -> list[Issue]:
        return [issue for issue in self.issues if issue.severity == "error"]


def as_text(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ("" if value is None else str(value).strip())


def normalise_header(value: Any) -> str:
    value = unicodedata.normalize("NFKC", as_text(value)).casefold()
    return re.sub(r"\s+", " ", value)


def slugify(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", "-", ascii_value).strip("-")


def is_blank(value: Any) -> bool:
    return not as_text(value)


def contains_placeholder(value: Any) -> bool:
    return PLACEHOLDER_TEXT in as_text(value).casefold()


def is_pale_yellow(cell: Any) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return False
    colour = fill.fgColor
    raw = colour.rgb if colour.type == "rgb" else ""
    return bool(raw and raw.upper().endswith(PALE_YELLOW))


def parse_price(value: Any) -> Decimal | None:
    cleaned = re.sub(r"[^0-9.,]", "", as_text(value)).replace(",", "")
    if not cleaned:
        return None
    try:
        price = Decimal(cleaned)
    except InvalidOperation:
        return None
    return price if price >= 0 else None


def format_price(value: Decimal) -> str:
    return "£" + format(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), ".2f")


def is_publishable_product(status: str, outbound_url: str) -> bool:
    """The only publication gate for either directory product type."""
    return status.casefold() == "available" and url_is_valid(outbound_url)


def boolean_active(value: Any) -> bool | None:
    normalised = as_text(value).upper()
    if normalised == "YES":
        return True
    if normalised == "NO":
        return False
    return None


def exact_int(value: Any) -> int | None:
    try:
        number = Decimal(as_text(value))
    except InvalidOperation:
        return None
    if number <= 0 or number != number.to_integral_value():
        return None
    return int(number)


def url_is_valid(value: Any) -> bool:
    return bool(URL_RE.match(as_text(value)))


def clean_list(value: Any) -> list[str]:
    return [item.strip() for item in as_text(value).split(",") if item.strip()]


def medium_facet(value: str) -> str:
    """Keep card copy intact while deriving a concise directory-filter label."""
    return re.sub(r"\s+patterns$", "", value, flags=re.IGNORECASE).strip() or value


def js_file(global_name: str, payload: dict[str, Any], source_name: str) -> str:
    generated = json.dumps(payload, ensure_ascii=False, indent=2)
    return (
        f"// Auto-generated by {SCRIPT_PATH} from {source_name}.\n"
        "// DO NOT EDIT MANUALLY — amend the workbook and regenerate.\n"
        f"window.{global_name} = {generated};\n"
    )


def maker_js_file(slug: str, payload: dict[str, Any], source_name: str) -> str:
    generated = json.dumps(payload, ensure_ascii=False, indent=2)
    return (
        f"// Auto-generated by {SCRIPT_PATH} from {source_name}.\n"
        "// DO NOT EDIT MANUALLY — amend the workbook and regenerate.\n"
        "window.HA_TO_MAKE_MAKERS = window.HA_TO_MAKE_MAKERS || {};\n"
        f"window.HA_TO_MAKE_MAKERS[{json.dumps(slug, ensure_ascii=False)}] = {generated};\n"
    )


def write_file(path: Path, content: str) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(content, encoding="utf-8")
    temporary.replace(path)


def display_header(options: Iterable[str]) -> str:
    return next(iter(options)).replace("\n", " ")


def header_map(worksheet: Any, row: int) -> dict[str, tuple[int, Any]]:
    result: dict[str, tuple[int, Any]] = {}
    for cell in worksheet[row]:
        key = normalise_header(cell.value)
        if key:
            result[key] = (cell.column, cell)
    return result


def locate_column(headers: dict[str, tuple[int, Any]], aliases: Iterable[str]) -> tuple[int, Any] | None:
    for alias in aliases:
        found = headers.get(normalise_header(alias))
        if found:
            return found
    return None


def field_cell(worksheet: Any, row: int, headers: dict[str, tuple[int, Any]], aliases: Iterable[str]) -> Any | None:
    located = locate_column(headers, aliases)
    return worksheet.cell(row=row, column=located[0]) if located else None


def validate_headers(audit: Audit, worksheet: Any, header_row: int, fields: dict[str, tuple[str, ...]], required: Iterable[str], label: str) -> dict[str, tuple[int, Any]]:
    headers = header_map(worksheet, header_row)
    for field in required:
        if not locate_column(headers, fields[field]):
            audit.add("fatal", worksheet.title, f"{header_row}:{header_row}", display_header(fields[field]), f"Required {label} header is missing or renamed.")
    return headers


def check_required_cell(audit: Audit, cell: Any | None, field: str, required: bool = True) -> str:
    if cell is None:
        return ""
    value = as_text(cell.value)
    if required and not value:
        audit.required_problem(cell.parent.title, cell.coordinate, field, "Required field is blank.")
    if required and contains_placeholder(value):
        audit.placeholder(cell.parent.title, cell.coordinate, field, "Contains unresolved PLACEHOLDER text.")
    if required and is_pale_yellow(cell):
        audit.placeholder(cell.parent.title, cell.coordinate, field, "Cell remains pale-yellow, the workbook’s unresolved-placeholder marker.")
    return value


def check_url_cell(audit: Audit, cell: Any | None, field: str, required: bool = True) -> str:
    value = check_required_cell(audit, cell, field, required)
    if value and not url_is_valid(value):
        audit.required_problem(cell.parent.title, cell.coordinate, field, "URL must be an absolute https:// URL.")
    return value


def profile_map(worksheet: Any) -> dict[str, Any]:
    entries: dict[str, Any] = {}
    for row in range(2, 16):
        label = normalise_header(worksheet.cell(row=row, column=1).value)
        if label:
            entries[label] = worksheet.cell(row=row, column=2)
    return entries


def profile_cell(entries: dict[str, Any], aliases: Iterable[str]) -> Any | None:
    for alias in aliases:
        cell = entries.get(normalise_header(alias))
        if cell:
            return cell
    return None


def locate_pattern_header_row(worksheet: Any) -> int | None:
    for row in range(1, min(worksheet.max_row, 100) + 1):
        labels = {normalise_header(cell.value) for cell in worksheet[row] if cell.value is not None}
        if "collection" in labels and any(label in labels for label in ("pattern title", "product title", "title")):
            return row
    return None


def row_has_values(worksheet: Any, row: int, columns: Iterable[int]) -> bool:
    return any(not is_blank(worksheet.cell(row=row, column=column).value) for column in columns)


def collection_groups(patterns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    ordering: dict[str, int] = {}
    for pattern in patterns:
        grouped[pattern["collection"]].append(pattern)
        ordering.setdefault(pattern["collection"], pattern["collectionOrder"])
    groups: list[dict[str, Any]] = []
    for collection in sorted(grouped, key=lambda item: (ordering[item], item.casefold())):
        products = sorted(grouped[collection], key=lambda item: (item["productOrder"], item["title"].casefold()))
        groups.append({"collection": collection, "patterns": products})
    return groups


def parse_contributor_sheet(audit: Audit, worksheet: Any, roster: dict[str, Any], kind: str) -> dict[str, Any]:
    """Parse one contributor tab while keeping staged product rows off the site.

    Product completeness is deliberately advisory. The publication contradiction is the
    only product-row condition that blocks a strict run: status=available without a
    usable outbound shop URL.
    """
    is_make = kind == "to-make"
    profile_fields = PROFILE_FIELDS if is_make else TO_KEEP_PROFILE_FIELDS
    required_profile_fields = REQUIRED_TO_MAKE_PROFILE_FIELDS if is_make else REQUIRED_TO_KEEP_PROFILE_FIELDS
    header_fields = TO_MAKE_PATTERN_HEADERS if is_make else TO_KEEP_PATTERN_HEADERS
    required_fields = REQUIRED_TO_MAKE_PATTERN_FIELDS if is_make else REQUIRED_TO_KEEP_PATTERN_FIELDS
    table_label = "Pattern" if is_make else "Product"
    image_url_fields = ("workbench_url", "lifestyle_url", "inside_url") if is_make else ("lifestyle_url", "flat_url", "detail_url")
    optional_fields = ("good_for_1", "good_for_2", "good_for_3", "coming_soon_label", "notes") if is_make else ("gift_for_1", "gift_for_2", "gift_for_3", "coming_soon_label", "notes")

    entries = profile_map(worksheet)
    profile: dict[str, str] = {}
    for field in required_profile_fields:
        cell = profile_cell(entries, profile_fields[field])
        if cell is None:
            audit.add("fatal", worksheet.title, "A2:B15", display_header(profile_fields[field]), "Required contributor-profile field is missing or renamed.")
            profile[field] = ""
        elif field.startswith("montage_") and field.endswith("_url"):
            profile[field] = check_url_cell(audit, cell, display_header(profile_fields[field]))
        else:
            profile[field] = check_required_cell(audit, cell, display_header(profile_fields[field]))

    header_row = locate_pattern_header_row(worksheet)
    if header_row is None:
        audit.add("fatal", worksheet.title, "1:100", f"{table_label} header row", f"Could not locate the {table_label.casefold()}-table header row.")
        return {"profile": profile, "patterns": [], "groups": []}
    headers = validate_headers(audit, worksheet, header_row, header_fields, required_fields, table_label.casefold())
    records: list[dict[str, Any]] = []
    seen_titles: dict[str, str] = {}
    header_columns = [column for column, _ in headers.values()]

    for row in range(header_row + 1, worksheet.max_row + 1):
        if not row_has_values(worksheet, row, header_columns):
            continue
        cells = {field: field_cell(worksheet, row, headers, header_fields[field]) for field in set(required_fields).union(optional_fields)}
        values = {field: as_text(cell.value) if cell is not None else "" for field, cell in cells.items()}
        status = values["status"].casefold()
        url_value = values["etsy_url"]
        publishable = is_publishable_product(status, url_value)
        title = values["title"]
        title_cell = cells["title"]
        status_cell = cells["status"]
        url_cell = cells["etsy_url"]

        if status not in {"available", "coming"}:
            audit.warning(worksheet.title, status_cell.coordinate if status_cell else str(row), "Status", "Not published: enter available only when this product has a usable outbound shop URL; otherwise use coming while staged.")
        elif status == "available" and not url_is_valid(url_value):
            audit.required_problem(worksheet.title, url_cell.coordinate if url_cell else str(row), "Outbound shop URL", "Available product has no usable absolute https:// shop URL and is withheld from publication.")
        elif not publishable:
            audit.warning(worksheet.title, status_cell.coordinate if status_cell else str(row), "Publication state", "Staged product is withheld from the site because status is not available.")

        # A publishable item can still be editorially incomplete. Surface it, but do not
        # convert normal in-progress sheet work into a strict-mode blocker.
        if publishable:
            for field in required_fields:
                if field in {"active", "etsy_url"}:
                    continue
                cell = cells[field]
                value = values[field]
                if not value:
                    audit.warning(worksheet.title, cell.coordinate if cell else str(row), display_header(header_fields[field]), "Published product is incomplete; the blank value is emitted as blank.")
                elif contains_placeholder(value) or (cell is not None and is_pale_yellow(cell)):
                    audit.warning(worksheet.title, cell.coordinate if cell else str(row), display_header(header_fields[field]), "Published product still contains an unresolved placeholder marker.")
            for field in image_url_fields:
                value = values[field]
                cell = cells[field]
                if value and not url_is_valid(value):
                    audit.warning(worksheet.title, cell.coordinate if cell else str(row), display_header(header_fields[field]), "Published product image URL should be an absolute https:// URL.")

        price_value = parse_price(values["price"])
        if price_value is None:
            if publishable:
                price_cell = cells["price"]
                audit.warning(worksheet.title, price_cell.coordinate if price_cell else str(row), display_header(header_fields["price"]), "Published product price is missing or invalid; no price is rendered.")
            price_display = ""
            price_number: float | None = None
        else:
            price_display = format_price(price_value)
            price_number = float(price_value)

        collection_order = exact_int(values["collection_order"])
        product_order = exact_int(values["product_order"])
        if publishable:
            for label, order, source_field in (("CollectionOrder", collection_order, "collection_order"), ("ProductOrder", product_order, "product_order")):
                if order is None:
                    cell = cells[source_field]
                    audit.warning(worksheet.title, cell.coordinate if cell else str(row), label, "Published product order should be a positive whole number; it will sort before ordered records.")

        title_slug = slugify(title)
        if publishable and not title_slug:
            audit.warning(worksheet.title, title_cell.coordinate if title_cell else str(row), table_label + " Title", "Published product has no title from which to derive a URL-safe slug.")
        if publishable and title_slug:
            existing = seen_titles.get(title_slug)
            if existing:
                audit.warning(worksheet.title, title_cell.coordinate if title_cell else str(row), table_label + " Title", f"Derived slug {title_slug!r} duplicates row {existing}.")
            seen_titles[title_slug] = str(row)

        record: dict[str, Any] = {
            "show": publishable,
            "slug": title_slug,
            "collection": values["collection"],
            "collectionOrder": collection_order or 0,
            "productOrder": product_order or 0,
            "title": title,
            "status": status,
            "format": values["format"],
            "price": price_display,
            "priceNumber": price_number,
            "etsyUrl": url_value,
            "meaning": values["description"],
            "comingSoonLabel": values["coming_soon_label"],
        }
        if is_make:
            record.update({
                "difficulty": values["difficulty"],
                "technique": values["technique"],
                "delivery": values["delivery"],
                "goodFor": [values[name] for name in ("good_for_1", "good_for_2", "good_for_3") if values[name]],
                "images": {"workbench": values["workbench_url"], "lifestyle": values["lifestyle_url"], "inside": values["inside_url"]},
                "imageAlts": {"workbench": values["workbench_alt"], "lifestyle": values["lifestyle_alt"], "inside": values["inside_alt"]},
            })
        else:
            record.update({
                "sizes": values["sizes"],
                "giftFor": [values[name] for name in ("gift_for_1", "gift_for_2", "gift_for_3") if values[name]],
                "images": {"lifestyle": values["lifestyle_url"], "flat": values["flat_url"], "detail": values["detail_url"]},
                "imageAlts": {"lifestyle": values["lifestyle_alt"], "flat": values["flat_alt"], "detail": values["detail_alt"]},
            })
        records.append(record)

    if profile["slug"] != roster["slug"]:
        audit.warning(worksheet.title, "B2", "Slug", f"Contributor-profile slug {profile['slug']!r} does not match roster slug {roster['slug']!r}.")
    if profile["display_name"] != roster["name"]:
        audit.warning(worksheet.title, "B3", "Display Name", "Contributor-profile display name does not match the roster.")
    if worksheet.title != roster["name"]:
        audit.warning(worksheet.title, "A1", "Contributor tab", "Contributor-tab title does not match the roster Display Name.")

    public_records = [record for record in records if record["show"]]
    groups = collection_groups(public_records)
    categories = [group["collection"] for group in groups]
    collection_count = len(categories)
    product_count = len(public_records)
    profile_count_label = "Stats — Patterns" if is_make else "Stats — Prints"
    if exact_int(profile["stats_collections"]) != collection_count:
        audit.warning(worksheet.title, "B7", "Stats — Collections", f"States {profile['stats_collections']!r}; the published site set contains {collection_count} collection(s). Generated statistics use the published set.")
    if exact_int(profile["stats_patterns"]) != product_count:
        audit.warning(worksheet.title, "B8", profile_count_label, f"States {profile['stats_patterns']!r}; the published site set contains {product_count} product(s). Generated statistics use the published set.")
    if set(roster["collections"]) != set(categories):
        audit.warning(worksheet.title, "B7", "Collections", "Roster collections do not match the published product collections; staged collections are intentionally not rendered.")

    prices = [Decimal(str(record["priceNumber"])) for record in public_records if record["priceNumber"] is not None]
    computed_from = format_price(min(prices)) if prices else ""
    if public_records and not prices:
        audit.warning(worksheet.title, "B9", "Stats — From Price", "Published products have no valid prices; no From price is rendered.")
    if computed_from:
        for source_name, source_value, source_cell in (("Roster Price From", roster["price_from"], roster["price_from_cell"]), ("Stats — From Price", profile["stats_from_price"], "B9")):
            if parse_price(source_value) != parse_price(computed_from):
                audit.warning(worksheet.title, source_cell if isinstance(source_cell, str) else source_cell.coordinate, source_name, f"States {source_value!r}; published product prices produce {computed_from}. Generated data uses the published minimum.")

    montage = [{"src": profile[f"montage_{index}_url"], "alt": profile[f"montage_{index}_alt"]} for index in range(1, 4)]
    profile_stats_key = "patterns" if is_make else "prints"
    return {
        "slug": roster["slug"], "active": roster["active"], "name": roster["name"], "buttonName": roster["button_name"], "artistNumber": roster["artist_number"],
        "badge": roster["badge"], "heroImage": roster["hero_image_url"], "heroAlt": roster["hero_image_alt"],
        "medium": roster["medium"], "filterMedium": medium_facet(roster["medium"]), "collections": roster["collections"],
        "priceFrom": computed_from, "feeling": roster["feeling"], "traits": roster["traits"], "cardUrl": roster["card_url"],
        "difficulty": roster.get("difficulty", ""), "technique": roster.get("technique", ""), "delivery": roster.get("delivery", ""),
        "patternCount": product_count,
        "profile": {"discipline": profile["discipline"], "bio": profile["bio"], "metaDescription": profile["meta_description"], "stats": {"collections": collection_count, profile_stats_key: product_count, "fromPrice": computed_from}, "montage": montage},
        "groups": groups,
    }


def parse_to_make_maker_sheet(audit: Audit, worksheet: Any, roster: dict[str, Any]) -> dict[str, Any]:
    return parse_contributor_sheet(audit, worksheet, roster, "to-make")


def parse_to_keep_maker_sheet(audit: Audit, worksheet: Any, roster: dict[str, Any]) -> dict[str, Any]:
    return parse_contributor_sheet(audit, worksheet, roster, "to-keep")

def parse_workbook(path: Path, kind: str, mode: str) -> tuple[Audit, list[dict[str, Any]]]:
    audit = Audit(mode)
    workbook = load_workbook(path, data_only=False)
    if kind == "auto":
        kind = "to-make" if "To Make Roster" in workbook.sheetnames else "to-keep" if "To Keep Roster" in workbook.sheetnames else "unknown"
    roster_name = "To Make Roster" if kind == "to-make" else "To Keep Roster"
    if kind == "unknown" or roster_name not in workbook.sheetnames:
        audit.add("fatal", "Workbook", "—", "Roster sheet", "Expected a sheet named To Make Roster or To Keep Roster.")
        return audit, []
    roster_sheet = workbook[roster_name]
    required_roster_fields = REQUIRED_TO_MAKE_ROSTER_FIELDS if kind == "to-make" else REQUIRED_TO_KEEP_ROSTER_FIELDS
    required_roster_headers = REQUIRED_ROSTER_HEADERS if kind == "to-make" else REQUIRED_TO_KEEP_ROSTER_HEADERS
    roster_headers = validate_headers(audit, roster_sheet, 1, TO_MAKE_ROSTER_HEADERS, required_roster_headers, "roster")
    rosters: list[dict[str, Any]] = []
    seen_slugs: dict[str, str] = {}
    header_columns = [column for column, _ in roster_headers.values()]
    for row in range(2, roster_sheet.max_row + 1):
        if not row_has_values(roster_sheet, row, header_columns):
            continue
        active_cell = field_cell(roster_sheet, row, roster_headers, TO_MAKE_ROSTER_HEADERS["active"])
        active = boolean_active(active_cell.value if active_cell else None)
        if active is None:
            if any(field_cell(roster_sheet, row, roster_headers, TO_MAKE_ROSTER_HEADERS[name]) and not is_blank(field_cell(roster_sheet, row, roster_headers, TO_MAKE_ROSTER_HEADERS[name]).value) for name in ("slug", "display_name")):
                audit.required_problem(roster_sheet.title, active_cell.coordinate if active_cell else str(row), "Active (YES/NO)", "Record has data but Active is not YES or NO.")
            continue
        values: dict[str, str] = {}
        cells: dict[str, Any] = {}
        for field in required_roster_fields:
            cell = field_cell(roster_sheet, row, roster_headers, TO_MAKE_ROSTER_HEADERS[field])
            cells[field] = cell
            if not active:
                # Active NO is an ordinary staging state. Retain its supplied material,
                # but never turn its incomplete fields into strict-run blockers.
                values[field] = as_text(cell.value) if cell is not None else ""
            elif field in {"card_url"}:
                values[field] = check_required_cell(audit, cell, display_header(TO_MAKE_ROSTER_HEADERS[field]))
            elif field == "hero_image_url":
                values[field] = check_url_cell(audit, cell, display_header(TO_MAKE_ROSTER_HEADERS[field]))
            else:
                values[field] = check_required_cell(audit, cell, display_header(TO_MAKE_ROSTER_HEADERS[field]))
        button_name_cell = field_cell(roster_sheet, row, roster_headers, TO_MAKE_ROSTER_HEADERS["button_name"])
        values["button_name"] = check_required_cell(audit, button_name_cell, "Button Name", required=False)
        artist_number_cell = field_cell(roster_sheet, row, roster_headers, TO_MAKE_ROSTER_HEADERS["artist_number"])
        artist_number_value = as_text(artist_number_cell.value) if artist_number_cell else ""
        artist_number = exact_int(artist_number_value)
        if active and artist_number is None:
            audit.blocking_error(
                roster_sheet.title,
                artist_number_cell.coordinate if artist_number_cell else str(row),
                "Artist Number",
                "Active contributor requires a unique positive whole Artist Number across both workbooks.",
            )
        elif not active and artist_number_value:
            audit.warning(
                roster_sheet.title,
                artist_number_cell.coordinate if artist_number_cell else str(row),
                "Artist Number",
                "Inactive contributor has an Artist Number; clear it until the contributor is active.",
            )
        if active and not url_is_valid(values["hero_image_url"]):
            audit.blocking_error(
                roster_sheet.title,
                cells["hero_image_url"].coordinate if cells["hero_image_url"] else str(row),
                "Hero Image URL",
                "Active contributor requires a usable absolute https:// Hero Image URL for directory and Home recent-artist cards.",
            )
        slug = values["slug"]
        if slug != slugify(slug):
            (audit.required_problem if active else audit.warning)(roster_sheet.title, cells["slug"].coordinate if cells["slug"] else str(row), "Slug", "Must be lower-case, URL-safe, and hyphen-separated.")
        if slug in seen_slugs:
            (audit.required_problem if active else audit.warning)(roster_sheet.title, cells["slug"].coordinate if cells["slug"] else str(row), "Slug", f"Duplicates roster row {seen_slugs[slug]}.")
        if slug:
            seen_slugs[slug] = str(row)
        if active and not values["card_url"].startswith("/to-make/") and kind == "to-make":
            audit.required_problem(roster_sheet.title, cells["card_url"].coordinate if cells["card_url"] else str(row), "Card URL", "To Make card URL must start with /to-make/.")
        if active and not values["card_url"].startswith("/to-keep/") and kind == "to-keep":
            audit.required_problem(roster_sheet.title, cells["card_url"].coordinate if cells["card_url"] else str(row), "Card URL", "To Keep card URL must start with /to-keep/.")
        rosters.append({
            "active": active,
            "slug": slug,
            "name": values["display_name"],
            "button_name": values["button_name"],
            "artist_number": artist_number,
            "badge": values["badge"],
            "medium": values["medium"],
            "collections": clean_list(values["collections"]),
            "price_from": values["price_from"],
            "price_from_cell": cells["price_from"],
            "difficulty": values.get("difficulty", ""),
            "technique": values.get("technique", ""),
            "delivery": values.get("delivery", ""),
            "feeling": values["feeling"],
            "traits": [values[name] for name in ("trait_1", "trait_2", "trait_3") if values[name]],
            "card_url": values["card_url"],
            "hero_image_url": values["hero_image_url"],
            "hero_image_alt": values["hero_image_alt"],
        })
        audit.roster_records.append({
            "kind": kind,
            "sheet": roster_sheet.title,
            "row": row,
            "slug": slug,
            "name": values["display_name"],
            "active": active,
            "artist_number": artist_number,
            "artist_number_value": artist_number_value,
            "artist_number_cell": artist_number_cell.coordinate if artist_number_cell else str(row),
        })

    makers: list[dict[str, Any]] = []
    for roster in rosters:
        if not roster["active"]:
            continue
        if roster["name"] not in workbook.sheetnames:
            audit.add("fatal", roster_sheet.title, roster["price_from_cell"].coordinate, "Maker tab", f"No sheet named {roster['name']!r} exists for active roster record {roster['slug']!r}.")
            continue
        parser = parse_to_make_maker_sheet if kind == "to-make" else parse_to_keep_maker_sheet
        makers.append(parser(audit, workbook[roster["name"]], roster))

    all_slugs: dict[str, str] = {}
    for maker in makers:
        for group in maker["groups"]:
            for pattern in group["patterns"]:
                duplicate = all_slugs.get(pattern["slug"])
                if duplicate:
                    audit.required_problem("Workbook", "—", "Pattern Title", f"Derived pattern slug {pattern['slug']!r} appears in both {duplicate} and {maker['slug']}.")
                all_slugs[pattern["slug"]] = maker["slug"]
    return audit, makers


def validate_artist_numbers(audit: Audit) -> None:
    """Validate the shared Artist Number series after both workbooks are parsed."""
    seen: dict[int, dict[str, Any]] = {}
    active_numbers: list[int] = []
    for record in audit.roster_records:
        number = record["artist_number"]
        raw = record["artist_number_value"]
        if raw and number is None:
            if record["active"]:
                audit.blocking_error(record["sheet"], record["artist_number_cell"], "Artist Number", "Artist Number must be a positive whole integer.")
            continue
        if number is None:
            continue
        if number in seen:
            first = seen[number]
            audit.blocking_error(
                record["sheet"],
                record["artist_number_cell"],
                "Artist Number",
                f"Artist Number {number} duplicates {first['sheet']}!{first['artist_number_cell']} ({first['name']}).",
            )
        else:
            seen[number] = record
        if record["active"]:
            active_numbers.append(number)
    if active_numbers:
        assigned = set(active_numbers)
        gaps = [str(value) for value in range(min(assigned), max(assigned) + 1) if value not in assigned]
        if gaps:
            audit.warning("Workbook", "—", "Artist Number", "Active Artist Number sequence has gap(s): " + ", ".join(gaps) + ". Gaps are allowed, but confirm that no number was unintentionally skipped.")


def recent_artists_payload(to_keep: list[dict[str, Any]], to_make: list[dict[str, Any]], source_names: str) -> dict[str, Any]:
    contributors: list[dict[str, Any]] = []
    for world, makers in (("TO KEEP", to_keep), ("TO MAKE", to_make)):
        for maker in makers:
            if maker["active"]:
                contributors.append({
                    "artistNumber": maker["artistNumber"],
                    "world": world,
                    "name": maker["name"],
                    "heroImage": maker["heroImage"],
                    "heroAlt": maker["heroAlt"],
                    "feeling": maker["feeling"],
                    "cardUrl": maker["cardUrl"],
                })
    contributors.sort(key=lambda item: item["artistNumber"], reverse=True)
    return {"generatedFrom": source_names, "artists": contributors[:4]}


def make_index_payload(makers: list[dict[str, Any]], source_name: str) -> dict[str, Any]:
    active = [maker for maker in makers if maker["active"]]
    pattern_count = sum(len(group["patterns"]) for maker in active for group in maker["groups"])
    medium_count = len({maker["medium"] for maker in active if maker["medium"]})
    return {
        "generatedFrom": source_name,
        "images": ASSETS,
        "navigation": NAVIGATION,
        "page": {
            "eyebrow": "To Make",
            "headingHtml": "Curated <em>Makers</em>",
            "intro": "A growing anthology of people who design things for other people to make. Patterns, templates and guides — chosen because making is one of the oldest ways there is of getting through something.",
            "stats": [
                {"value": len(active), "label": "FOUNDING MAKERS"},
                {"value": pattern_count, "label": "PATTERNS"},
                {"value": medium_count, "label": "MEDIUM SO FAR" if medium_count == 1 else "MEDIUMS SO FAR"},
            ],
        },
        "filters": {
            "medium": ["All mediums"] + sorted({maker["filterMedium"] for maker in active if maker["filterMedium"]}),
            "level": ["All levels", "Beginner", "Improver"],
            "delivery": ["All delivery", "Instant download", "Posted to you"],
        },
        "recruitment": {
            "eyebrow": "FOUNDING MAKERS",
            "heading": "Do you design things for other people to make?",
            "body": "Patterns, templates, guides — any medium. Free for the founding year, no commission ever. I just need some images, a bit of a description about you, and some links.",
            "linkLabel": "See how it works for artists →",
            "linkUrl": "/for-artists",
        },
        "collective": COLLECTIVE,
        "footer": FOOTER,
        "makers": [
            {
                key: maker[key]
                for key in ("slug", "active", "name", "buttonName", "artistNumber", "badge", "heroImage", "heroAlt", "medium", "filterMedium", "collections", "priceFrom", "difficulty", "technique", "delivery", "feeling", "traits", "cardUrl", "patternCount")
            }
            for maker in active
        ],
    }


def keep_index_payload(makers: list[dict[str, Any]], source_name: str) -> dict[str, Any]:
    active = [maker for maker in makers if maker["active"]]
    published_artists = [maker for maker in active if any(group["patterns"] for group in maker["groups"])]
    published_works = [
        product
        for maker in published_artists
        for group in maker["groups"]
        for product in group["patterns"]
    ]
    published_collections = {product["collection"] for product in published_works if product["collection"]}
    return {
        "generatedFrom": source_name,
        "images": ASSETS,
        "navigation": NAVIGATION,
        "page": {
            "eyebrow": "To Keep",
            "headingHtml": "Curated <em>Artists</em>",
            "intro": "A growing anthology of artists whose work carries meaning — pieces to lift you, anchor you, or simply stay with you. Every artist is here because their work holds a feeling worth keeping.",
            "stats": [
                {"value": len(published_artists), "label": "FOUNDING ARTISTS", "suffix": "of the first 50"},
                {"value": len(published_works), "label": "WORKS"},
                {"value": len(published_collections), "label": "COLLECTIONS"},
            ],
        },
        "filters": {"medium": ["All"] + sorted({maker["medium"] for maker in active if maker["medium"]}), "price": ["All", "Under £20", "£20–£50", "£50+"]},
        "recruitment": {"eyebrow": "FOUNDING ARTISTS", "heading": "Your work could be here.", "body": "Free for the founding year. No commission ever. I just need some images, a bit of a description about you, and some links.", "linkLabel": "See how it works for artists →", "linkUrl": "/for-artists"},
        "collective": COLLECTIVE,
        "footer": FOOTER,
        "artists": [{key: maker[key] for key in ("slug", "active", "name", "buttonName", "artistNumber", "badge", "heroImage", "heroAlt", "medium", "filterMedium", "collections", "priceFrom", "feeling", "traits", "cardUrl")} for maker in active],
    }


TO_KEEP_SHARED_CHROME = {
    # These constants deliberately preserve the current To Keep profile chrome without
    # depending on a previous generated file.
    "images": {
        "logo": "https://images.squarespace-cdn.com/content/6a258894c750534b28845855/1f3bfb43-558e-4262-b2e2-d4e3b56bd77e/01-the-hope-anthology.jpg?content-type=image%2Fjpeg",
        "star": "https://images.squarespace-cdn.com/content/6a258894c750534b28845855/0956e10d-38c6-4fd7-8e06-7f2579fb75d6/collections-footer-star-hope-anthology.png?content-type=image%2Fpng",
    },
    "navigation": [
        {"label": "To Keep", "url": "/to-keep"}, {"label": "To Make", "url": "/to-make"},
        {"label": "The Story", "url": "/story"}, {"label": "Collaborate", "url": "/for-artists"},
        {"label": "Collective", "url": "/collective"},
    ],
    "collective": {"heading": "Something worth being part of.", "body": "Join the Collective for new print notes, quiet launches, and gentle studio updates.", "buttonLabel": "Join the Collective", "buttonUrl": "/collective"},
    "footer": {"instagramUrl": "https://www.instagram.com/thehopeanthology", "privacyUrl": "/privacy-policy", "accessibilityUrl": "/accessibility", "sellingUrl": "/why-we-sell-this-way", "copyright": "© The Hope Anthology 2026"},
}


def keep_profile_record(maker: dict[str, Any]) -> dict[str, Any]:
    profile = maker["profile"]
    products = [product for group in maker["groups"] for product in group["patterns"]]
    collections = [group["collection"] for group in maker["groups"]]
    return {
        "slug": maker["slug"],
        "path": "/to-keep/" + maker["slug"],
        "world": "To Keep",
        "eyebrow": "TO KEEP · ANTHOLOGY ARTIST",
        "artistName": maker["name"],
        "artistTitle": profile["discipline"],
        "discipline": profile["discipline"],
        "bio": profile["bio"],
        "metaTitle": maker["name"] + " — " + profile["discipline"] + " | The Hope Anthology",
        "metaDescription": profile["metaDescription"],
        "montage": profile["montage"],
        "stats": profile["stats"],
        "breadcrumbs": [{"label": "To Keep", "url": "/to-keep"}, {"label": maker["name"]}],
        "filterOptions": ["All"] + collections,
        "gridEyebrow": maker["name"].upper() + " — ALL WORK",
        "products": products,
    }


def keep_profiles_payload(makers: list[dict[str, Any]], source_name: str) -> dict[str, Any]:
    return {"generatedFrom": source_name, **TO_KEEP_SHARED_CHROME, "collections": {maker["slug"]: keep_profile_record(maker) for maker in makers if maker["active"]}}


def maker_payload(maker: dict[str, Any], source_name: str) -> dict[str, Any]:
    return {"generatedFrom": source_name, "images": ASSETS, "navigation": NAVIGATION, "collective": COLLECTIVE, "footer": FOOTER, "maker": maker}

def report_markdown(source: Path, mode: str, makers: list[dict[str, Any]], audit: Audit, output_paths: list[Path]) -> str:
    active_patterns = sum(len(group["patterns"]) for maker in makers for group in maker["groups"])
    strict_blockers = len(audit.errors) + (len(audit.placeholders) if mode == "permissive" else 0)
    lines = [
        "# Hope Anthology Directory Workbook Generation Report",
        "",
        f"**Source workbook:** `{source.name}`  ",
        f"**Generator:** `{SCRIPT_PATH}`  ",
        f"**Mode:** `{mode}`  ",
        f"**Outcome:** `{'not generated' if audit.fatal or (mode == 'strict' and audit.errors) else 'generated'}`",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
        f"| Active makers | {len(makers)} |",
        f"| Published products/patterns | {active_patterns} |",
        f"| Generated content files | {len(output_paths)} |",
        f"| Fatal structural issues | {len(audit.fatal)} |",
        f"| Strict-mode blockers | {strict_blockers} |",
        f"| Unresolved placeholders | {len(audit.placeholders)} |",
        "",
        "## Generated files",
        "",
    ]
    if output_paths:
        lines += [f"- `{path.name}`" for path in output_paths]
    else:
        lines.append("- None — structural validation stopped output.")
    lines += ["", "## Unresolved placeholder report", ""]
    if audit.placeholders:
        lines += ["| Sheet | Cell | Field | Finding |", "| --- | --- | --- | --- |"]
        for issue in audit.placeholders:
            lines.append(f"| {issue.sheet} | {issue.cell} | {issue.field} | {issue.message} |")
    else:
        lines.append("No unresolved placeholder markers were found.")
    lines += ["", "## Validation findings", ""]
    non_placeholder = [issue for issue in audit.issues if issue not in audit.placeholders]
    if non_placeholder:
        lines += ["| Severity | Sheet | Cell | Field | Finding |", "| --- | --- | --- | --- | --- |"]
        for issue in non_placeholder:
            lines.append(f"| {issue.severity.upper()} | {issue.sheet} | {issue.cell} | {issue.field} | {issue.message} |")
    else:
        lines.append("No non-placeholder validation findings.")
    lines += [
        "",
        "## Mode behaviour",
        "",
        "Product rows are staged by default. A product is emitted only when Status is available and its outbound shop URL is a usable absolute https:// URL. In strict mode, the only product-row error is an available product without that URL; incomplete or coming rows are reported as warnings and withheld. Missing required sheets, headers, contributor tabs, or product-table headers remain fatal.",
        "",
    ]
    return "\n".join(lines)


def output_pairs(kind: str, makers: list[dict[str, Any]], source_name: str, output_dir: Path) -> list[tuple[Path, str]]:
    if kind == "to-make":
        outputs: list[tuple[Path, str]] = [
            (output_dir / "content.to-make.js", js_file("HA_TO_MAKE_CONTENT", make_index_payload(makers, source_name), source_name)),
        ]
        for maker in makers:
            outputs.append((output_dir / f"content.to-make.{maker['slug']}.js", maker_js_file(maker["slug"], maker_payload(maker, source_name), source_name)))
        return outputs
    # The To Keep workbook drives both the parent directory and the profile registry
    # consumed by ha-artist-page.js. No merge with old content occurs.
    return [
        (output_dir / "content.to-keep.js", js_file("HA_TO_KEEP_CONTENT", keep_index_payload(makers, source_name), source_name)),
        (output_dir / "content.keep-collections.js", js_file("HA_KEEP_COLLECTIONS_CONTENT", keep_profiles_payload(makers, source_name), source_name)),
    ]


def finish_generation(source_name: str, mode: str, makers: list[dict[str, Any]], audit: Audit, outputs: list[tuple[Path, str]], report_path: Path) -> int:
    output_paths: list[Path] = []
    may_generate = not audit.fatal and not audit.blocked and not (mode == "strict" and audit.errors)
    if may_generate:
        for target, content in outputs:
            write_file(target, content)
            output_paths.append(target)
    write_file(report_path, report_markdown(Path(source_name), mode, makers, audit, output_paths))
    for issue in audit.issues:
        print(f"{issue.severity.upper()}: {issue.sheet}!{issue.cell} [{issue.field}] {issue.message}")
    print(f"REPORT: {report_path}")
    for target in output_paths:
        print(f"GENERATED: {target}")
    if audit.fatal:
        return 2
    if audit.blocked or (mode == "strict" and audit.errors):
        return 1
    return 0


def build_directory(source: Path, kind: str, output_dir: Path, mode: str, report_path: Path) -> int:
    audit, makers = parse_workbook(source, kind, mode)
    validate_artist_numbers(audit)
    return finish_generation(source.name, mode, makers, audit, output_pairs(kind, makers, source.name, output_dir), report_path)


def build_all_directories(to_keep_source: Path, to_make_source: Path, output_dir: Path, mode: str, report_path: Path) -> int:
    """Generate both directory families and the Home feed as one validation-gated run."""
    keep_audit, keep_makers = parse_workbook(to_keep_source, "to-keep", mode)
    make_audit, make_makers = parse_workbook(to_make_source, "to-make", mode)
    audit = Audit(mode)
    audit.issues = keep_audit.issues + make_audit.issues
    audit.placeholders = keep_audit.placeholders + make_audit.placeholders
    audit.roster_records = keep_audit.roster_records + make_audit.roster_records
    audit.blocked = keep_audit.blocked or make_audit.blocked
    validate_artist_numbers(audit)
    source_name = to_keep_source.name + " + " + to_make_source.name
    outputs = output_pairs("to-keep", keep_makers, to_keep_source.name, output_dir)
    outputs += output_pairs("to-make", make_makers, to_make_source.name, output_dir)
    outputs.append((output_dir / "content.home-recent-artists.js", js_file("HA_HOME_RECENT_ARTISTS", recent_artists_payload(keep_makers, make_makers, source_name), source_name)))
    return finish_generation(source_name, mode, keep_makers + make_makers, audit, outputs, report_path)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Hope Anthology directory content from one or both sibling workbooks.")
    parser.add_argument("workbook", type=Path, nargs="?", help="Optional single .xlsx workbook for legacy one-directory generation.")
    parser.add_argument("--kind", choices=("to-make", "to-keep", "auto"), default="auto", help="Kind for a legacy one-workbook run.")
    parser.add_argument("--to-keep-workbook", type=Path, help="To Keep sibling workbook for the normal whole-Anthology generation run.")
    parser.add_argument("--to-make-workbook", type=Path, help="To Make sibling workbook for the normal whole-Anthology generation run.")
    parser.add_argument("--mode", choices=("permissive", "strict"), default="strict", help="Permissive reports normal staged material; strict blocks product publication contradictions.")
    parser.add_argument("--output-dir", type=Path, default=Path.cwd(), help="Directory for generated content files.")
    parser.add_argument("--report", type=Path, help="Path for the Markdown validation report.")
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    if args.to_keep_workbook or args.to_make_workbook:
        if args.workbook or not args.to_keep_workbook or not args.to_make_workbook:
            print("ERROR: the whole-Anthology run requires both --to-keep-workbook and --to-make-workbook, with no positional workbook.", file=sys.stderr)
            return 2
        if not args.to_keep_workbook.is_file() or not args.to_make_workbook.is_file():
            print("ERROR: a specified workbook was not found.", file=sys.stderr)
            return 2
        report = (args.report or output_dir / "directory-generation-report.md").resolve()
        return build_all_directories(args.to_keep_workbook.resolve(), args.to_make_workbook.resolve(), output_dir, args.mode, report)
    if not args.workbook or not args.workbook.is_file():
        print("ERROR: provide one workbook or both sibling workbook options.", file=sys.stderr)
        return 2
    kind = args.kind
    if kind == "auto":
        probe = load_workbook(args.workbook, read_only=True, data_only=False)
        kind = "to-make" if "To Make Roster" in probe.sheetnames else "to-keep" if "To Keep Roster" in probe.sheetnames else "unknown"
    if kind == "unknown":
        print("ERROR: workbook does not contain To Make Roster or To Keep Roster.", file=sys.stderr)
        return 2
    report_name = "to-make-generation-report.md" if kind == "to-make" else "to-keep-generation-report.md"
    report = (args.report or output_dir / report_name).resolve()
    return build_directory(args.workbook.resolve(), kind, output_dir, args.mode, report)


if __name__ == "__main__":
    raise SystemExit(main())
